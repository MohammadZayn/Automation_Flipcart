import pytest
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

# Path to the Excel file
excel_file_path = r"C:\Users\local_\Downloads\Excel_data.xlsx"


# Function to read data from Excel file using openpyxl
def read_excel_data(file_path, sheet_name='Sheet1'):
    workbook = load_workbook(file_path)
    sheet = workbook[sheet_name]

    # Extracting data from Excel, skipping the header
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row + ("",))  # Adding an empty slot for the result

    return data, workbook, sheet


# Loading data from Excel
test_data, workbook, sheet = read_excel_data(excel_file_path)


@pytest.mark.parametrize("row_index, id, username, password, result",
                         [(index + 2, *row) for index, row in enumerate(test_data)])
def test_login(row_index, id, username, password, result):
    driver = webdriver.Chrome()  # Use the appropriate WebDriver for your browser
    try:
        driver.get("https://example.com/login")  # Replace with your target website URL

        # Example login automation logic
        driver.find_element(By.ID, "username").send_keys(username)
        driver.find_element(By.ID, "password").send_keys(password)
        driver.find_element(By.ID, "loginBtn").click()

        # Implement your verification logic here (e.g., checking if login was successful)
        if "Welcome" in driver.page_source:
            sheet.cell(row=row_index, column=5, value="Pass")  # Column 5 corresponds to the "Result" column
            print(f"Test {id} with Username: {username} passed.")
        else:
            sheet.cell(row=row_index, column=5, value="Fail")
            print(f"Test {id} with Username: {username} failed.")
    except Exception as e:
        sheet.cell(row=row_index, column=5, value="Fail")
        print(f"Error during test execution for {username}: {e}")
    finally:
        driver.quit()


# Save the test results back to the Excel file after all tests
def pytest_sessionfinish(session, exitstatus):
    workbook.save(excel_file_path)
    print(f"Test results updated in '{excel_file_path}'")
